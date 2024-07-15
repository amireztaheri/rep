import os
import streamlit as st
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
import persian
import requests
import convertapi

# تنظیمات ConvertAPI
convertapi.api_secret = 'l7xcVHa66mDMKYVG'

# اتصال به دیتابیس
DATABASE = 'database.db'

def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

# تبدیل عدد به حروف فارسی
def convert_number_to_farsi_words(number):
    units = ["", "یک", "دو", "سه", "چهار", "پنج", "شش", "هفت", "هشت", "نه"]
    teens = ["ده", "یازده", "دوازده", "سیزده", "چهارده", "پانزده", "شانزده", "هفده", "هجده", "نوزده"]
    tens = ["", "", "بیست", "سی", "چهل", "پنجاه", "شصت", "هفتاد", "هشتاد", "نود"]
    hundreds = ["", "صد", "دویست", "سیصد", "چهارصد", "پانصد", "ششصد", "هفتصد", "هشتصد", "نهصد"]

    if number == 0:
        return "صفر"

    parts = []

    if number >= 100:
        parts.append(hundreds[number // 100])
        number %= 100

    if 10 <= number < 20:
        parts.append(teens[number - 10])
    else:
        parts.append(tens[number // 10])
        number %= 10
        parts.append(units[number])

    return " ".join([part for part in parts if part])

# به‌روزرسانی فایل اکسل
def update_excel(file_path, student):
    wb = load_workbook(file_path)
    ws = wb.active

    font = Font(name='B Karim', size=12)
    cells = ['U2', 'V3', 'U4', 'V5', 'V6', 'U7', 'Z4', 'Z5', 'Z6', 'Z7',
             'K2', 'K3', 'J4', 'J5', 'J6', 'J7', 'G6', 'F7', 'H7', 'F8']
    inputs = [student['name'], student['last_name'], student['father_name'], student['ssn'], student['birth_place'],
              student['birth_date'], student['national_id'], student['student_type'], student['birth_location'],
              student['student_code'], student['province'], student['region'], student['school'], student['branch'],
              student['field'], student['standard_code'], student['transcript'], student['school_year'], student['term'],
              student['school_type']]
    
    for cell, input_value in zip(cells, inputs):
        ws[cell].value = input_value
        ws[cell].font = font

    save_path = os.path.join(os.getcwd(), 'updated_madrak.xlsx')
    wb.save(save_path)
    
    # Convert to PDF
    result = convertapi.convert('pdf', {
        'File': save_path
    }, from_format='xlsx')
    pdf_path = save_path.replace('.xlsx', '.pdf')
    result.save_files(pdf_path)
    return pdf_path

# ارسال فایل به تلگرام
def send_to_telegram(pdf_path):
    url = f"https://api.telegram.org/botYOUR_BOT_API/sendDocument"
    files = {'document': open(pdf_path, 'rb')}
    data = {'chat_id': 'YOUR_CHAT_ID'}
    response = requests.post(url, files=files, data=data)
    if response.status_code == 200:
        st.success("PDF file has been sent via Telegram.")
    else:
        st.error(f"Failed to send PDF: {response.status_code}\n{response.text}")

# رابط کاربری Streamlit
st.title("مدیریت دانش‌آموزان")

conn = get_db_connection()
students = conn.execute('SELECT * FROM students').fetchall()
conn.close()

df_students = pd.DataFrame(students, columns=students[0].keys())

st.dataframe(df_students)

selected_student_id = st.selectbox("انتخاب دانش‌آموز", df_students['id'])
selected_student = df_students[df_students['id'] == selected_student_id].iloc[0]

uploaded_file = st.file_uploader("فایل اکسل را آپلود کنید", type=['xlsx'])

if st.button("به‌روزرسانی فایل اکسل"):
    if uploaded_file is not None:
        file_path = os.path.join("data", "madrak.xlsx")
        with open(file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        pdf_path = update_excel(file_path, selected_student)
        send_to_telegram(pdf_path)
        st.success("فایل اکسل به‌روزرسانی شد و به PDF تبدیل شد.")

